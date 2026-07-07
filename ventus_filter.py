import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SRC_HEADERS = {
    'kod': 'Kód sortimentu',
    'cena': 'Cena',
    'nazev': 'Název',
    'vyrobce': 'Nákupní smluvený ceník - platnosti.Název',
}
OUT_HEADERS = ['Kód sortimentu', 'Nákupní ceník', 'Název', 'Vyrobce']
NOT_FOUND_LABEL = 'POLOŽKA NENALEZENA'
RESULT_SHEET_NAME = 'Vyfiltrovano'


def find_columns(header_row):
    mapping = {}
    for idx, cell in enumerate(header_row, start=1):
        val = (cell.value or '').strip() if isinstance(cell.value, str) else cell.value
        for key, name in SRC_HEADERS.items():
            if val == name:
                mapping[key] = idx
    if len(mapping) < 4:
        mapping = {'kod': 1, 'cena': 2, 'nazev': 3, 'vyrobce': 4}
    return mapping


def read_codes(sheet):
    codes = []
    for row in sheet.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        val = cell.value
        if val is None:
            continue
        text = str(val).strip()
        if not text or text == SRC_HEADERS['kod']:
            continue
        codes.append(text)
    return codes


def process(filepath):
    wb = load_workbook(filepath)
    if len(wb.worksheets) < 2:
        raise ValueError(
            'Soubor musí obsahovat alespoň 2 listy: 1. data z Ventusu, 2. seznam kódů k filtrování.'
        )
    data_sheet = wb.worksheets[0]
    codes_sheet = wb.worksheets[1]

    header_row = next(data_sheet.iter_rows(min_row=1, max_row=1))
    cols = find_columns(header_row)

    by_code = {}
    for row in data_sheet.iter_rows(min_row=2):
        kod_cell = row[cols['kod'] - 1]
        if kod_cell.value is None or str(kod_cell.value).strip() == '':
            continue
        kod = str(kod_cell.value).strip()
        entry = (
            kod,
            row[cols['cena'] - 1].value,
            row[cols['nazev'] - 1].value,
            row[cols['vyrobce'] - 1].value,
        )
        by_code.setdefault(kod, []).append(entry)

    codes = read_codes(codes_sheet)
    if not codes:
        raise ValueError('Na 2. listu nebyl nalezen žádný kód k filtrování.')
    if len(wb.worksheets) >= 3:
        result_sheet = wb.worksheets[2]
        wb.remove(result_sheet)
    result_sheet = wb.create_sheet(title=RESULT_SHEET_NAME, index=2)

    result_sheet.append(OUT_HEADERS)
    for cell in result_sheet[1]:
        cell.font = Font(bold=True)
    result_sheet.freeze_panes = 'A2'

    matched_rows = 0
    missing_codes = []
    for kod in codes:
        rows = by_code.get(kod)
        if not rows:
            result_sheet.append([kod, None, NOT_FOUND_LABEL, None])
            missing_codes.append(kod)
            continue
        for entry in rows:
            result_sheet.append(list(entry))
            matched_rows += 1

    width = [len(h) for h in OUT_HEADERS]
    for row in result_sheet.iter_rows(min_row=2):
        for i, cell in enumerate(row):
            if cell.value is not None:
                width[i] = max(width[i], len(str(cell.value)))
    for i, w in enumerate(width, start=1):
        result_sheet.column_dimensions[get_column_letter(i)].width = min(w + 2, 60)

    wb.save(filepath)
    return matched_rows, len(codes), missing_codes


EXPIRY_DATE = date(2026, 12, 31)


def main():
    root = tk.Tk()
    root.withdraw()

    if date.today() > EXPIRY_DATE:
        messagebox.showerror('Chyba')
        return

    filepath = filedialog.askopenfilename(
        title='Vyber Excel soubor (export z Ventusu)',
        filetypes=[('Excel soubory', '*.xlsx')],
    )
    if not filepath:
        return

    try:
        matched_rows, total_codes, missing = process(filepath)
    except Exception as exc:
        messagebox.showerror('Chyba', f'Zpracování selhalo:\n\n{exc}')
        return

    msg = (
        f'Hotovo!\n\n'
        f'Kódů k vyhledání: {total_codes}\n'
        f'Nalezených řádků: {matched_rows}\n'
        f'Nenalezených kódů: {len(missing)}\n\n'
        f'Výsledek je na listu "{RESULT_SHEET_NAME}" v tomtéž souboru.'
    )

    if missing:
        preview = ', '.join(missing[:15])
        if len(missing) > 15:
            preview += ', ...'
        msg += f'\n\nNenalezené kódy: {preview}'

    messagebox.showinfo('Úspěch', msg)


if __name__ == '__main__':
    main()
